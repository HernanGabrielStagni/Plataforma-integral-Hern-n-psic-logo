from flask import Flask, request, jsonify, send_file, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import os
import sys
import webbrowser
import threading

app = Flask(__name__, static_folder=".", static_url_path="")

MESES_ES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

# Feriados nacionales argentinos 2025-2027
FERIADOS_NACIONALES = {
    2025: [
        (1, 1), (3, 3), (3, 4), (3, 24), (4, 2), (4, 17), (4, 18),
        (5, 1), (5, 2), (5, 25), (6, 16), (6, 20), (7, 9),
        (8, 17), (10, 12), (11, 20), (11, 24), (12, 8), (12, 25)
    ],
    2026: [
        (1, 1), (2, 16), (2, 17), (3, 24), (4, 2), (4, 2), (4, 3),
        (5, 1), (5, 25), (6, 15), (6, 20), (7, 9),
        (8, 17), (10, 12), (11, 20), (11, 23), (12, 8), (12, 25)
    ],
    2027: [
        (1, 1), (2, 8), (2, 9), (3, 24), (3, 25), (3, 26),
        (4, 2), (5, 1), (5, 24), (5, 25), (6, 20), (6, 21),
        (7, 9), (8, 16), (10, 11), (11, 20), (11, 22), (12, 8), (12, 25)
    ],
}

# Feriado local: 11 de noviembre (Día de Río Cuarto)
FERIADO_LOCAL = (11, 11)


@app.route("/")
def index():
    return send_from_directory(".", "index.html")


@app.route("/api/feriados")
def get_feriados():
    year = int(request.args.get("year", 2025))
    month = int(request.args.get("month", 1))

    nacionales = []
    if year in FERIADOS_NACIONALES:
        nacionales = [d for m, d in FERIADOS_NACIONALES[year] if m == month]

    local = []
    if FERIADO_LOCAL[0] == month:
        local = [FERIADO_LOCAL[1]]

    return jsonify({"nacionales": nacionales, "local": local})


MESES_CORTO = [
    "", "ene", "feb", "mar", "abr", "may", "jun",
    "jul", "ago", "sep", "oct", "nov", "dic"
]


@app.route("/api/generar", methods=["POST"])
def generar_excel():
    import calendar
    import datetime

    data = request.json
    year = int(data["year"])
    month = int(data["month"])
    feriados = set(int(d) for d in data.get("feriados", []))

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # Columnas: A=Fecha D, B=M, C=A, D=Hora Ingreso, E=Hora Egreso, F=Hs, G=Firma, H=Obs
    col_widths = {
        "A": 12, "B": 7, "C": 7, "D": 20,
        "E": 20, "F": 9, "G": 26, "H": 18
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Estilos
    azul_oscuro = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    verde_claro = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    naranja_claro = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    blanco_bold = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    blanco_font = Font(name="Arial", color="FFFFFF", size=10)
    normal_font = Font(name="Arial", size=9)
    bold_font = Font(name="Arial", bold=True, size=9)
    small_font = Font(name="Arial", size=8)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=False)

    def style_range(row, col_start, col_end, font=None, fill=None, alignment=None, border=True):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=c)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = thin_border

    total_days = calendar.monthrange(year, month)[1]
    mes_corto = MESES_CORTO[month]
    year_short = str(year)[2:]  # "26" para 2026
    mes_nombre = MESES_ES[month]

    # =========================================================
    # FILA 1: Titulo + ASPURC
    # =========================================================
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="PLANILLA DE ASIST. DIARIA - REHABILITACIÓN -Acomp. Terapeutico")
    c.font = blanco_bold
    c.fill = azul_oscuro
    c.alignment = center
    ws.merge_cells("G1:H1")
    c = ws.cell(row=1, column=7, value="ASPURC")
    c.font = blanco_bold
    c.fill = azul_oscuro
    c.alignment = center
    style_range(1, 1, 8, border=True)

    # =========================================================
    # FILA 2: Prestador | Especialidad | Matrícula
    # =========================================================
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:B2")
    ws.cell(row=2, column=1, value="Prestador:").font = bold_font
    ws.cell(row=2, column=1).alignment = left_align
    ws.merge_cells("C2:D2")
    ws.cell(row=2, column=3, value="Lic. Stagni Hernán Gabriel").font = normal_font
    ws.cell(row=2, column=3).alignment = left_align
    ws.merge_cells("E2:F2")
    ws.cell(row=2, column=5, value="Especialidad: Psicología").font = normal_font
    ws.cell(row=2, column=5).alignment = left_align
    ws.merge_cells("G2:H2")
    ws.cell(row=2, column=7, value="Matrícula: 11919").font = normal_font
    ws.cell(row=2, column=7).alignment = right_align
    style_range(2, 1, 8)

    # =========================================================
    # FILA 3: CUIT | Dirección
    # =========================================================
    ws.row_dimensions[3].height = 18
    ws.merge_cells("A3:B3")
    ws.cell(row=3, column=1, value="CUIT:").font = bold_font
    ws.cell(row=3, column=1).alignment = left_align
    ws.merge_cells("C3:D3")
    ws.cell(row=3, column=3, value="20.24521983.1").font = normal_font
    ws.cell(row=3, column=3).alignment = left_align
    ws.merge_cells("E3:H3")
    ws.cell(row=3, column=5, value="Dirección del lugar de trabajo: Vilelas 2223 - Río Cuarto").font = normal_font
    ws.cell(row=3, column=5).alignment = left_align
    style_range(3, 1, 8)

    # =========================================================
    # FILA 4: Mes | Período | mes | Modalidad
    # =========================================================
    ws.row_dimensions[4].height = 18
    ws.merge_cells("A4:B4")
    ws.cell(row=4, column=1, value="Mes:").font = bold_font
    ws.cell(row=4, column=1).alignment = left_align
    ws.merge_cells("C4:D4")
    ws.cell(row=4, column=3, value=f"{mes_corto}-{year_short}").font = normal_font
    ws.cell(row=4, column=3).alignment = left_align
    ws.cell(row=4, column=5, value=f"Período  01 al {total_days}").font = normal_font
    ws.cell(row=4, column=5).alignment = left_align
    ws.cell(row=4, column=6, value=f"{mes_corto}-{year_short}").font = normal_font
    ws.cell(row=4, column=6).alignment = left_align
    ws.merge_cells("G4:H4")
    ws.cell(row=4, column=7, value="Modalidad de Acomp. : Mixta").font = normal_font
    ws.cell(row=4, column=7).alignment = left_align
    style_range(4, 1, 8)

    # =========================================================
    # FILA 5: Nombre y Apellido + N.Afiliado (fondo amarillo)
    # =========================================================
    ws.row_dimensions[5].height = 22
    ws.merge_cells("A5:H5")
    c = ws.cell(row=5, column=1,
                value="Nombre y Apellido:  Morales Ferreira, Alma Ángela.....N.Afiliado:  54939539")
    c.font = bold_font
    c.fill = amarillo
    c.alignment = left_align
    style_range(5, 1, 8, fill=amarillo)

    # =========================================================
    # FILA 6: Encabezados de tabla
    # =========================================================
    ws.row_dimensions[6].height = 22
    ws.merge_cells("A6:C6")
    ws.cell(row=6, column=1, value="Fecha D/M/A").font = bold_font
    ws.cell(row=6, column=4, value="Hora de Ingreso").font = bold_font
    ws.cell(row=6, column=5, value="Hora de Egreso").font = bold_font
    ws.cell(row=6, column=6, value="Hs.").font = bold_font
    ws.cell(row=6, column=7, value="Firma Titular/Respons.").font = bold_font
    ws.cell(row=6, column=8, value="Obs./Dx.").font = bold_font
    style_range(6, 1, 8, alignment=center)

    # =========================================================
    # FILAS DE DATOS (fila 7 en adelante)
    # =========================================================
    data_start_row = 7

    for day in range(1, total_days + 1):
        dt = datetime.date(year, month, day)
        dow = dt.weekday()  # 0=lunes, 6=domingo
        es_feriado = day in feriados
        es_sabado = dow == 5
        es_domingo = dow == 6
        es_viernes = dow == 4

        ingreso = ""
        egreso = ""
        horas = ""
        obs = ""

        if es_feriado:
            obs = "FERIADO"
        elif es_domingo:
            pass
        elif es_viernes:
            pass
        elif es_sabado:
            ingreso = "19:00"
            egreso = "21:30"
            horas = 2.50
        elif dow in (0, 1, 2, 3):
            ingreso = "15:00"
            egreso = "17:30"
            horas = 2.50

        r = data_start_row + day - 1

        ws.cell(row=r, column=1, value=day).font = normal_font
        ws.cell(row=r, column=2, value=f"{month:02d}").font = normal_font
        ws.cell(row=r, column=3, value=year_short).font = normal_font
        ws.cell(row=r, column=4, value=ingreso).font = normal_font
        ws.cell(row=r, column=5, value=egreso).font = normal_font

        horas_cell = ws.cell(row=r, column=6)
        if isinstance(horas, (int, float)) and horas > 0:
            horas_cell.value = horas
            horas_cell.number_format = "0.00"
        else:
            horas_cell.value = ""
        horas_cell.font = normal_font

        ws.cell(row=r, column=7, value="").font = normal_font
        ws.cell(row=r, column=8, value=obs).font = normal_font

        # Colores de fondo
        fill = None
        if es_feriado or es_domingo:
            fill = naranja_claro
        elif es_sabado:
            fill = verde_claro

        for col in range(1, 9):
            c = ws.cell(row=r, column=col)
            c.alignment = center
            c.border = thin_border
            if fill:
                c.fill = fill

    # =========================================================
    # FILA DE TOTAL
    # =========================================================
    total_row = data_start_row + total_days
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    total_label = ws.cell(row=total_row, column=1, value="TOTAL HORAS")
    total_label.font = Font(name="Arial", bold=True, size=10)
    total_label.alignment = right_align

    total_cell = ws.cell(row=total_row, column=6)
    total_cell.value = f"=SUM(F{data_start_row}:F{data_start_row + total_days - 1})"
    total_cell.font = Font(name="Arial", bold=True, size=10)
    total_cell.number_format = "0.00"
    total_cell.alignment = center

    for col in range(1, 9):
        ws.cell(row=total_row, column=col).border = thin_border

    # =========================================================
    # NOTAS AL PIE
    # =========================================================
    notas = [
        "1. La presente planilla debe ser firmada por el paciente o su representante legal en cada sesión.",
        "2. El prestador debe presentar esta planilla junto con la factura correspondiente para su liquidación.",
        "3. La acreditación del prestador debe estar vigente al momento de la prestación."
    ]
    nota_row = total_row + 2
    for i, nota in enumerate(notas):
        ws.merge_cells(start_row=nota_row + i, start_column=1, end_row=nota_row + i, end_column=8)
        c = ws.cell(row=nota_row + i, column=1, value=nota)
        c.font = Font(name="Arial", size=8, italic=True)
        c.alignment = left_align

    # =========================================================
    # CONFIGURAR IMPRESIÓN (A4, horizontal)
    # =========================================================
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # --- GUARDAR ---
    filename = f"Asistencia Alma {mes_nombre} {year}.xlsx"

    # Guardar en la carpeta Descargas estándar de Windows
    downloads_dir = os.path.expanduser("~\\Downloads")
    os.makedirs(downloads_dir, exist_ok=True)
    filepath = os.path.join(downloads_dir, filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)


def open_browser():
    webbrowser.open("http://127.0.0.1:5050")


if __name__ == "__main__":
    threading.Timer(1.5, open_browser).start()
    app.run(host="127.0.0.1", port=5050, debug=False)
